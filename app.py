import streamlit as st
import pandas as pd
import re
import io
from pypdf import PdfReader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurazione della pagina Streamlit
st.set_page_config(
    page_title="DMI Trasporti - Elaboratore Liste",
    page_icon="🚚",
    layout="wide"
)

# Stile CSS personalizzato per un look professionale ed elegante
st.markdown("""
    <style>
    .main-title {
        font-size: 28pt;
        font-weight: bold;
        color: #1E3A8A;
        margin-bottom: 5px;
    }
    .subtitle {
        font-size: 12pt;
        color: #555555;
        margin-bottom: 25px;
    }
    .metric-box {
        background-color: #F3F4F6;
        padding: 15px;
        border-radius: 8px;
        border-left: 5px solid #1E3A8A;
        margin-bottom: 15px;
    }
    .metric-title {
        font-size: 10pt;
        color: #6B7280;
        text-transform: uppercase;
        font-weight: bold;
    }
    .metric-value {
        font-size: 18pt;
        font-weight: bold;
        color: #111827;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">🚚 DMI Trasporti - Gestore Liste di Carico</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Applicazione ottimizzata per l\'elaborazione automatica, l\'aggregazione e l\'esportazione blindata in Excel senza dati fantasma.</div>', unsafe_allow_html=True)

# --- FUNZIONI DI PULIZIA E PARSING ---

def pulisci_volume(valore):
    if pd.isna(valore) or valore is None:
        return 0.0
    val_str = str(valore).strip().upper()
    val_str = val_str.replace("MC", "").replace("SI", "").replace("NO", "").strip()
    
    # Correzione del doppio punto Nippon Express (es. 125.20.00 -> 125.20)
    if val_str.count('.') > 1:
        parti = val_str.split('.')
        val_str = parti[0] + '.' + parti[1]
        
    val_str = re.sub(r'[^0-9.]', '', val_str)
    try:
        return float(val_str) if val_str else 0.0
    except ValueError:
        return 0.0

def accorcia_nome(nome):
    if pd.isna(nome) or not nome:
        return ""
    nome = str(nome).upper().strip()
    # Rimozi ragioni sociali ed estensioni fastidiose
    for parola in ["SRL", "SPA", "SAS", "S.R.L.", "S.P.A.", "S.A.S.", "GRUP", "SOLUTIONS", "FURNITURE", "ARREDAMENTI", "ITALIA", "DI", "E C.", "S.N.C.", "SNC"]:
        nome = re.sub(r'\b' + parola + r'\b', '', nome)
    nome = re.sub(r'\s+', ' ', nome).strip()
    return nome[:15] # Taglio netto a 15 caratteri per pulizia grafica

def accorcia_localita(loc):
    if pd.isna(loc) or not loc:
        return ""
    loc = str(loc).upper().strip()
    loc = re.sub(r'[^A-Z\s]', '', loc).strip()
    return loc[:4] # Massimo 4 lettere maiuscole

def estrai_dati_da_pdf(file_oggetto, nome_autista):
    righe_estratte = []
    try:
        reader = PdfReader(file_oggetto)
        for pagina in reader.pages:
            testo = pagina.extract_text()
            if not testo:
                continue
                
            linee = testo.split('\n')
            for linea in linee:
                linea_str = linea.strip()
                parti = linea_str.split()
                if len(parti) >= 4:
                    try:
                        numeri = [p for p in parti if re.match(r'^[0-9.]+$', p)]
                        if len(numeri) >= 2:
                            imballi_val = int(float(numeri[-2]))
                            vol_val = pulisci_volume(numeri[-1])
                            
                            if "TOTALE" in linea_str.upper() or "TOTAL" in linea_str.upper():
                                continue
                                
                            testi_soli = [p for p in parti if not re.match(r'^[0-9.]+$', p)]
                            
                            if len(testi_soli) >= 3:
                                mittente = testi_soli[0]
                                localita = testi_soli[-1]
                                destinatario = " ".join(testi_soli[1:-1])
                                
                                righe_estratte.append({
                                    "AUTISTA": nome_autista,
                                    "MITTENTE": accorcia_nome(mittente),
                                    "DESTINATARIO": accorcia_nome(destinatario),
                                    "LOCALITÀ": accorcia_localita(localita),
                                    "IMBALLI": imballi_val,
                                    "VOLUME": vol_val
                                })
                    except:
                        continue
    except Exception as e:
        st.error(f"Errore nella lettura del file di {nome_autista}: {e}")
    return righe_estratte

# --- INTERFACCIA UTENTE E LOGICA DI SESSIONE ---

# Svuotiamo la cache persistente di Streamlit che causava l'errore dei file vecchi
st.cache_data.clear()

# Caricamento file multipli
uploaded_files = st.file_uploader(
    "1. Trascina qui i file PDF delle liste di carico (Puoi caricarne più di uno contemporaneamente)",
    type=["pdf"],
    accept_multiple_files=True,
    key="uploader_liste"
)

if not uploaded_files:
    st.info("👋 Nessun file caricato al momento. Trascina uno o più PDF per iniziare l'elaborazione istantanea.")
    st.stop()

# COSTRUZIONE DEL DATABASE CORRENTE (SOLO dai file attualmente presenti nel widget)
tutte_le_righe = []

for file_pdf in uploaded_files:
    nome_file = file_pdf.name
    nome_autista_rilevato = nome_file.split('.')[0].replace("lista_", "").replace("lista", "").strip().upper()
    
    dati_file = estrai_dati_da_pdf(file_pdf, nome_autista_rilevato)
    tutte_le_righe.extend(dati_file)

if len(tutte_le_righe) == 0:
    st.warning("⚠️ Il sistema ha letto i PDF ma la struttura del testo richiede una verifica. Di seguito trovi la tabella editabile: puoi incollare o correggere i dati direttamente nelle celle!")
    primo_autista = uploaded_files[0].name.split('.')[0].upper()
    tutte_le_righe.append({
        "AUTISTA": primo_autista, "MITTENTE": "ESEMPIO", "DESTINATARIO": "CLIENTE", "LOCALITÀ": "PALE", "IMBALLI": 1, "VOLUME": 0.45
    })

df_raw = pd.DataFrame(tutte_le_righe)

st.markdown("### 2. Revisione e Modifica Dati in Tempo Reale")
st.write("Puoi fare doppio clic su qualsiasi cella della tabella qui sotto per correggere nomi, località o volumi letti dal PDF prima di generare l'Excel definitivo.")

df_modificato = st.data_editor(
    df_raw,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "AUTISTA": st.column_config.TextColumn("Autista (Nome File)"),
        "MITTENTE": st.column_config.TextColumn("Mittente"),
        "DESTINATARIO": st.column_config.TextColumn("Destinatario"),
        "LOCALITÀ": st.column_config.TextColumn("Località (4 Lett.)", max_chars=4),
        "IMBALLI": st.column_config.NumberColumn("Imballi (Colli)", min_value=0, step=1),
        "VOLUME": st.column_config.NumberColumn("Volume (mc)", min_value=0.0, format="%.4f", step=0.0001),
    },
    key="editor_dati"
)

# Adeguamento e pulizia post-modifica manuale
df_modificato['VOLUME'] = df_modificato['VOLUME'].apply(puliisci_volume)
df_modificato['IMBALLI'] = pd.to_numeric(df_modificato['IMBALLI'], errors='coerce').fillna(0).astype(int)
df_modificato['MITTENTE'] = df_modificato['MITTENTE'].apply(accorcia_nome)
df_modificato['DESTINATARIO'] = df_modificato['DESTINATARIO'].apply(accorcia_nome)
df_modificato['LOCALITÀ'] = df_modificato['LOCALITÀ'].apply(accorcia_localita)

# --- FASE DI ELABORAZIONE, ALGORITMO E AGGREGAZIONE ---

# 1. Raggruppamento automatico (Group By) per unire le consegne identiche dello stesso autista
df_aggregato = df_modificato.groupby(
    ['AUTISTA', 'MITTENTE', 'DESTINATARIO', 'LOCALITÀ'], 
    as_index=False
).agg({
    'IMBALLI': 'sum',
    'VOLUME': 'sum'
})

# 2. Controllo Globale del Destinatario per l'Eccezione delle Consegne Minime
volumi_totali_destinatari = df_aggregato.groupby('DESTINATARIO')['VOLUME'].sum().to_dict()

def determina_minima(riga):
    dest = riga['DESTINATARIO']
    vol_riga = riga['VOLUME']
    vol_totale_cliente = volumi_totali_destinatari.get(dest, 0.0)
    
    if vol_totale_cliente >= 1.0:
        return ""
    elif vol_riga < 1.0:
        return "MINIMA"
    return ""

df_aggregato['CONSEGNA MINIMA'] = df_aggregato.apply(determina_minima, axis=1)

# --- CALCOLO METRICHE GENERALI ---
vol_totale_generale = df_aggregato['VOLUME'].sum()
df_solo_minime = df_aggregato[df_aggregato['CONSEGNA MINIMA'] == "MINIMA"]
vol_totale_minime = df_solo_minime['VOLUME'].sum()
num_totale_minime = len(df_solo_minime)

st.markdown("### 3. Indicatori Globali del Carico Attuale")
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown(f'<div class="metric-box"><div class="metric-title">Volume Totale Generale</div><div class="metric-value">{vol_totale_generale:.4f} mc</div></div>', unsafe_allow_html=True)
with col2:
    st.markdown(f'<div class="metric-box"><div class="metric-title">Volume Totale Consegne Minime</div><div class="metric-value">{vol_totale_minime:.4f} mc</div></div>', unsafe_allow_html=True)
with col3:
    st.markdown(f'<div class="metric-box"><div class="metric-title">Numero Consegne Minime</div><div class="metric-value">{num_totale_minime} spedizioni</div></div>', unsafe_allow_html=True)

# --- CREAZIONE DEL FILE EXCEL CON OPENPYXL (STRUTTURA BLINDATA) ---

def genera_excel_professionale(df_dati):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riepilogo Consegne"
    ws.views.sheetView[0].showGridLines = True
    
    font_titolo_autista = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    fill_titolo_autista = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    font_normale = Font(name="Calibri", size=11)
    font_minima = Font(name="Calibri", size=11, bold=True, color="990000")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    border_sottile = Side(style='thin', color='D1D5DB')
    bordo_cella = Border(left=border_sottile, right=border_sottile, top=border_sottile, bottom=border_sottile)
    
    headers = ["MITTENTE", "DESTINATARIO", "LOCALITÀ", "IMBALLI", "VOLUME (mc)", "CONSEGNA MINIMA"]
    riga_corrente = 1
    
    autisti = df_dati['AUTISTA'].unique()
    
    for autista in autisti:
        # Riga separatrice dell'Autista unita centralmente su tutte e 6 le colonne
        ws.merge_cells(start_row=riga_corrente, start_column=1, end_row=riga_corrente, end_column=6)
        cella_autista = ws.cell(row=riga_corrente, column=1, value=f"🚚 SPEDIZIONE AUTISTA: {autista}")
        cella_autista.font = font_titolo_autista
        cella_autista.fill = fill_titolo_autista
        cella_autista.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[riga_corrente].height = 26
        
        for col in range(1, 7):
            ws.cell(row=riga_corrente, column=col).border = bordo_cella
            
        riga_corrente += 1
        
        # Scrittura Intestazione Colonne fisse sotto l'autista
        for col_idx, header in enumerate(headers, 1):
            cella = ws.cell(row=riga_corrente, column=col_idx, value=header)
            cella.font = font_header
            cella.fill = fill_header
            cella.alignment = Alignment(horizontal="left" if col_idx <= 3 else "right", vertical="center")
            cella.border = bordo_cella
        ws.row_dimensions[riga_corrente].height = 20
        riga_corrente += 1
        
        df_autista = df_dati[df_dati['AUTISTA'] == autista]
        
        for idx, row in df_autista.iterrows():
            c_mitt = ws.cell(row=riga_corrente, column=1, value=row['MITTENTE'])
            c_dest = ws.cell(row=riga_corrente, column=2, value=row['DESTINATARIO'])
            c_loc = ws.cell(row=riga_corrente, column=3, value=row['LOCALITÀ'])
            c_imb = ws.cell(row=riga_corrente, column=4, value=row['IMBALLI'])
            c_vol = ws.cell(row=riga_corrente, column=5, value=row['VOLUME'])
            c_min = ws.cell(row=riga_corrente, column=6, value=row['CONSEGNA MINIMA'])
            
            c_imb.number_format = '#,##0'
            c_vol.number_format = '0.0000'
            
            c_mitt.alignment = Alignment(horizontal="left", vertical="center")
            c_dest.alignment = Alignment(horizontal="left", vertical="center")
            c_loc.alignment = Alignment(horizontal="center", vertical="center")
            c_imb.alignment = Alignment(horizontal="right", vertical="center")
            c_vol.alignment = Alignment(horizontal="right", vertical="center")
            c_min.alignment = Alignment(horizontal="center", vertical="center")
            
            for c in [c_mitt, c_dest, c_loc, c_imb, c_vol, c_min]:
                c.font = font_normale
                c.border = bordo_cella
                if idx % 2 == 0:
                    c.fill = fill_zebra
                    
            if row['CONSEGNA MINIMA'] == "MINIMA":
                c_min.font = font_minima
                
            ws.row_dimensions[riga_corrente].height = 18
            riga_corrente += 1
            
        riga_corrente += 1
        
    # --- RIGHE TOTALI E SUMMARY MATEMATICO FINALE ---
    riga_corrente += 1
    ws.cell(row=riga_corrente, column=1, value="RIEPILOGO LOGISTICO GENERALE").font = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
    riga_corrente += 1
    
    ws.cell(row=riga_corrente, column=1, value="Volume Totale Generale:").font = Font(name="Calibri", size=11, bold=True)
    激_tot_vol = ws.cell(row=riga_corrente, column=2, value=vol_totale_generale)
    激_tot_vol.number_format = '0.0000'
    激_tot_vol.font = Font(name="Calibri", size=11, bold=True)
    riga_corrente += 1
    
    ws.cell(row=riga_corrente, column=1, value="Volume Totale Consegne Minime:").font = Font(name="Calibri", size=11, bold=True)
    激_min_vol = ws.cell(row=riga_corrente, column=2, value=vol_totale_minime)
    激_min_vol.number_format = '0.0000'
    激_min_vol.font = Font(name="Calibri", size=11, bold=True)
    riga_corrente += 1
    
    ws.cell(row=riga_corrente, column=1, value="Numero Totale Consegne Minime:").font = Font(name="Calibri", size=11, bold=True)
    激_num_min = ws.cell(row=riga_corrente, column=2, value=num_totale_minime)
    激_num_min.number_format = '#,##0'
    激_num_min.font = Font(name="Calibri", size=11, bold=True)
    
    # Auto-adattamento larghezza colonne per evitare testi troncati (salvaguardando la colonna A dalle celle unite)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ws.merged_cells:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

excel_file = genera_excel_professionale(df_aggregato)

st.markdown("### 4. Esportazione File Excel")
st.download_button(
    label="📥 Scarica File Excel Elaborato (.xlsx)",
    data=excel_file,
    file_name="Riepilogo_Carichi_DMI.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.success("✨ Tabella data-editor e file Excel pronti con successo! Nessun residuo storico in background.")
